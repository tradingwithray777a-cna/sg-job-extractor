from __future__ import annotations

from datetime import datetime
from typing import Dict, List, Tuple
import re

from src.scoring import compute_relevance, closing_passed

from src.connectors.foundit import FounditConnector
from src.connectors.fastjobs import FastJobsConnector
from src.connectors.mycareersfuture import MyCareersFutureConnector

try:
    from src.excel_writer import write_excel
except Exception:
    write_excel = None


def _fallback_write_excel(jobs: List[Dict], notes: Dict, out_path: str) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    JOBS_COLS = [
        "Job title available",
        "employer",
        "job post url link",
        "job post from what source",
        "date job post was posted",
        "application closing date",
        "key job requirement",
        "estimated salary",
        "job full-time or part-time",
        "Relevance score",
        "Closing date passed? (Y/N)",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Jobs"
    ws.append(JOBS_COLS)
    ws.freeze_panes = "A2"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for col in range(1, len(JOBS_COLS) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font

    for r in jobs:
        ws.append([r.get(c, "") for c in JOBS_COLS])

    # hyperlink URL column (C)
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=3)
        if isinstance(cell.value, str) and cell.value.startswith("http"):
            cell.hyperlink = cell.value
            cell.font = Font(color="0000EE", underline="single")

    end_col = get_column_letter(len(JOBS_COLS))
    end_row = ws.max_row
    tab = Table(displayName="JobsTable", ref=f"A1:{end_col}{end_row}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(tab)

    ws2 = wb.create_sheet("Notes")
    ws2.append(["Item", "Value"])
    ws2.freeze_panes = "A2"
    ws2.column_dimensions["A"].width = 42
    ws2.column_dimensions["B"].width = 120
    ws2["A1"].font = Font(bold=True)
    ws2["B1"].font = Font(bold=True)
    for k, v in (notes or {}).items():
        ws2.append([str(k), str(v)])

    wb.save(out_path)
    return out_path


def _norm(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip().lower())


class KeywordSets:
    def __init__(
        self,
        target_role: str,
        core_keywords: List[str],
        adjacent_titles: List[str],
        nearby_titles: List[str],
        exclude_keywords: List[str],
    ):
        self.target_role = target_role
        self.core_keywords = core_keywords
        self.adjacent_titles = adjacent_titles
        self.nearby_titles = nearby_titles
        self.exclude_keywords = exclude_keywords


def build_keyword_sets(target_role: str) -> KeywordSets:
    tr = (target_role or "").strip()

    core = [tr]
    adjacent = [
        f"{tr} Executive",
        f"{tr} Officer",
        f"{tr} Specialist",
        f"Senior {tr}",
        f"Assistant {tr}",
        f"{tr} Coordinator",
        f"{tr} Manager",
    ]
    nearby = ["Programme Executive", "Stakeholder Management", "Partnerships Executive", "Community Engagement"]
    exclude = []

    # Add small “helpful” defaults for Community Partnership
    trn = _norm(tr)
    if "community" in trn and "partnership" in trn:
        core = [tr, "community", "partnership", "stakeholder", "engagement", "outreach", "volunteer"]
        adjacent = [
            "Community Partnerships Executive",
            "Partnerships Executive",
            "Partnership Officer",
            "Community Engagement Executive",
            "Stakeholder Engagement Executive",
            "Community Outreach Executive",
            "Programme Executive (Partnership)",
            "Corporate Partnerships Executive",
            "Strategic Partnerships Executive",
            "Partnership Development Executive",
        ]
        nearby = [
            "Programme Executive",
            "Programme Coordinator",
            "CSR Executive",
            "Corporate Relations Executive",
            "Business Development Executive",
            "Events Executive",
            "Volunteer Management Executive",
            "Community Manager",
        ]

    def dedupe(lst: List[str], cap: int) -> List[str]:
        out, seen = [], set()
        for x in lst:
            xn = _norm(x)
            if xn and xn not in seen:
                out.append(x)
                seen.add(xn)
            if len(out) >= cap:
                break
        return out

    return KeywordSets(
        target_role=tr,
        core_keywords=dedupe(core, 10),
        adjacent_titles=dedupe(adjacent, 20),
        nearby_titles=dedupe(nearby, 20),
        exclude_keywords=dedupe(exclude, 10),
    )


# ✅ PORTALS WIRED HERE
CONNECTORS = {
    "MyCareersFuture": MyCareersFutureConnector(),
    "Foundit": FounditConnector(),
    "FastJobs": FastJobsConnector(),
}


def build_queries(target_role: str, adjacent_titles: List[str], core_keywords: List[str]) -> List[Tuple[str, str]]:
    exact = (target_role, "Exact")

    adj = (adjacent_titles[0] if adjacent_titles else target_role, "Adjacent")

    cores = [c for c in (core_keywords or []) if _norm(c) != _norm(target_role)]
    skill = (f"{target_role} {cores[0]} {cores[1]}" if len(cores) >= 2 else target_role, "Skill-based")

    return [exact, adj, skill]


def run_search(
    target_role: str,
    posted_within_days: int,
    selected_portals: List[str],
    max_final: int = 100,
    raw_cap: int = 200,
    out_path: str = "output.xlsx",
) -> str:
    ks = build_keyword_sets(target_role)
    queries = build_queries(target_role, ks.adjacent_titles, ks.core_keywords)

    raw_rows: List[Dict] = []
    portal_stats: Dict[str, Dict[str, int]] = {}

    # Keep threshold moderate. If you want “some data no matter what”, reduce to 30.
    min_relevance = 30

    # Collect up to raw_cap rows across portals
    for portal in selected_portals:
        conn = CONNECTORS.get(portal)
        if not conn:
            portal_stats[portal] = {"returned": 0, "kept_after_score": 0}
            continue

        returned_total = 0

        for q, qtype in queries:
            try:
                jobs = conn.search(q, limit=80)
            except Exception:
                jobs = []

            returned_total += len(jobs)

            for j in jobs:
                raw_rows.append({
                    "Job title available": getattr(j, "title", "") or "Not stated",
                    "employer": getattr(j, "employer", "") or "Not stated",
                    "job post url link": getattr(j, "url", "") or "",
                    "job post from what source": portal,
                    "date job post was posted": getattr(j, "posted_date", "") or "Unverified",
                    "application closing date": getattr(j, "closing_date", "") or "Not stated",
                    "key job requirement": getattr(j, "requirements", "") or "• Not stated",
                    "estimated salary": getattr(j, "salary", "") or "Not stated",
                    "job full-time or part-time": getattr(j, "job_type", "") or "Not stated",
                })

                if len(raw_rows) >= raw_cap:
                    break

            if len(raw_rows) >= raw_cap:
                break

        portal_stats[portal] = {"returned": returned_total, "kept_after_score": 0}
        if len(raw_rows) >= raw_cap:
            break

    # Deduplicate (title+employer)
    def key(r): return (_norm(r["Job title available"]), _norm(r["employer"]))
    best = {}
    for r in raw_rows:
        k = key(r)
        if k not in best:
            best[k] = r
    deduped = list(best.values())

    # Score + closing
    today = datetime.now().date()
    for r in deduped:
        r["Relevance score"] = compute_relevance(r, ks.target_role, ks.adjacent_titles, ks.nearby_titles)
        r["Closing date passed? (Y/N)"] = closing_passed(r.get("application closing date", "Not stated"), today=today)

    kept = [r for r in deduped if int(r.get("Relevance score", 0)) >= min_relevance]

    # Portal kept stats
    for p in portal_stats:
        portal_stats[p]["kept_after_score"] = sum(1 for r in kept if r.get("job post from what source") == p)

    # Sort: relevance desc; verified posted date above unverified
    def sort_key(r):
        ver = 0 if r.get("date job post was posted") == "Unverified" else 1
        return (-int(r.get("Relevance score", 0)), -ver)

    kept.sort(key=sort_key)
    final = kept[:max_final]

    notes = {
        "Search date/time (SG time)": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "TARGET_ROLE": ks.target_role,
        "Queries used": "; ".join([f"{qt}:{q}" for q, qt in queries]),
        "Portals selected": ", ".join(selected_portals),
        "Portal stats (returned vs kept_after_score)": "; ".join(
            [f"{p}: {portal_stats[p]['returned']} returned, {portal_stats[p]['kept_after_score']} kept"
             for p in portal_stats]
        ),
        "Counts (raw → dedupe → kept_after_score → final)": f"{len(raw_rows)} → {len(deduped)} → {len(kept)} → {len(final)}",
        "Relevance threshold used": str(min_relevance),
        "Note": "If Foundit shows 0 returned, it is likely blocked/JS-rendered on Streamlit Cloud. MyCareersFuture should provide results for professional roles.",
    }

    if write_excel:
        return write_excel(final, notes, out_path)

    return _fallback_write_excel(final, notes, out_path)
