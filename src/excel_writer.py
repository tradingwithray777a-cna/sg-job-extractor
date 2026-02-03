from __future__ import annotations

from typing import List, Dict
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


JOBS_COLS = [
    "Job title available",
    "employer",
    "job post url link",
    "job post from what source",
    "date job post was posted",
    "application closing date",
    "key job requirement",
    "estimated salary",
    "job full-time or part-time",
    "Relevance score",
    "Closing date passed? (Y/N)",
]


def write_excel(jobs: List[Dict], notes: Dict, out_path: str) -> str:
    """
    Writes:
      - Sheet 1: Jobs (as an Excel table, filters, frozen header, hyperlink URLs)
      - Sheet 2: Notes
    Returns out_path.
    """
    # Build Jobs dataframe with exact column order
    df = pd.DataFrame(jobs)
    for c in JOBS_COLS:
        if c not in df.columns:
            df[c] = ""
    df = df.reindex(columns=JOBS_COLS)

    # Notes dataframe
    df_notes = pd.DataFrame([{"Item": k, "Value": v} for k, v in (notes or {}).items()])

    # Write initial workbook
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Jobs", index=False)
        df_notes.to_excel(writer, sheet_name="Notes", index=False)

    wb = load_workbook(out_path)

    # ---------------- Jobs sheet formatting ----------------
    ws = wb["Jobs"]
    ws.freeze_panes = "A2"

    # Header styling
    header_fill = PatternFill("solid", fgColor="1F4E79")  # deep blue
    header_font = Font(color="FFFFFF", bold=True)
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font

    # Hyperlinks on URL column (3rd column)
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=3)
        if isinstance(cell.value, str) and cell.value.startswith("http"):
            cell.hyperlink = cell.value
            cell.font = Font(color="0000EE", underline="single")

    # Add Excel table (filters are included)
    end_col = get_column_letter(ws.max_column)
    end_row = ws.max_row
    tab = Table(displayName="JobsTable", ref=f"A1:{end_col}{end_row}")
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showRowStripes=True,
        showColumnStripes=False
    )
    ws.add_table(tab)

    # Auto column width (simple heuristic, first 200 rows)
    for col in range(1, ws.max_column + 1):
        max_len = 10
        for r in range(1, min(ws.max_row, 200) + 1):
            v = ws.cell(row=r, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = min(60, max_len + 2)

    # ---------------- Notes sheet formatting ----------------
    ws2 = wb["Notes"]
    ws2.freeze_panes = "A2"
    ws2.column_dimensions["A"].width = 38
    ws2.column_dimensions["B"].width = 120

    # Save
    wb.save(out_path)
    return out_path
